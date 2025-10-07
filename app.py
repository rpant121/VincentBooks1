import streamlit as st
import requests
import pandas as pd
import random
import hashlib
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# --- Streamlit page config ---
st.set_page_config(page_title="Fable Book Explorer", layout="wide")
st.title("Fable Book Explorer")
st.caption("Browse and analyze books from a Fable reading list (powered by Fable's public API).")

# --- Constants ---
BASE_URL = "https://api.fable.co/api/v2/users/0c031026-9f1f-4a02-889c-79d2bdb11781/book_lists/33f803be-3e8d-4ed6-bd19-79a330d2bb32/books"
HEADERS = {"User-Agent": "Mozilla/5.0"}


# --- Utility: Deterministic genre color ---
def color_from_genre(genre: str) -> str:
    """Generate a consistent hex color for each genre based on its name."""
    if not genre:
        return "#888888"
    h = int(hashlib.sha1(genre.encode("utf-8")).hexdigest(), 16)
    random.seed(h)
    return f"#{random.randint(0, 0xFFFFFF):06x}"


@st.cache_data(show_spinner=False)
def fetch_all_books():
    """Fetch all paginated books from the Fable API and return a DataFrame."""
    all_books = []
    url = BASE_URL

    while url:
        resp = requests.get(url, headers=HEADERS)
        resp.raise_for_status()
        data = resp.json()
        results = data.get("results", [])
        if not results:
            break

        for item in results:
            book = item.get("book", {})
            all_books.append(book)

        url = data.get("next")

    # --- Format into DataFrame ---
    rows = []
    for b in all_books:
        authors = b.get("authors", [])
        genres = b.get("genres", [])
        finished_at = b.get("finished_reading_at")

        finished_dt = None
        if finished_at:
            try:
                finished_dt = datetime.fromisoformat(finished_at.replace("Z", "+00:00"))
            except Exception:
                pass

        rows.append({
            "title": b.get("title"),
            "subtitle": b.get("subtitle"),
            "author": authors[0]["name"] if authors else None,
            "genre": genres[0]["name"] if genres else "Unknown",
            "pages": b.get("page_count"),
            "isbn": b.get("isbn"),
            "published_date": b.get("published_date"),
            "imprint": b.get("imprint"),
            "cover_image": b.get("cover_image"),
            "book_url": f"https://fable.co/book/{b.get('id')}" if b.get("id") else None,
            "started_reading": b.get("started_reading_at"),
            "finished_reading": finished_at,
            "finished_datetime": finished_dt,
        })
    return pd.DataFrame(rows)


# --- Excel writer helper ---
def df_to_excel_with_colors(df: pd.DataFrame, genre_colors: dict) -> BytesIO:
    """Write a DataFrame to an Excel file with genre-based color highlights."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Books"

    # Header row
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        genre = row.get("genre", "Unknown")
        color = genre_colors.get(genre, "#FFFFFF").replace("#", "")
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# --- Genre legend helper ---
def display_genre_legend(genre_colors: dict):
    """Display a horizontal color legend for genres."""
    st.markdown("#### Genre Color Legend")
    cols = st.columns(len(genre_colors))
    for i, (genre, color) in enumerate(genre_colors.items()):
        with cols[i]:
            st.markdown(
                f"<div style='background-color:{color}; padding:6px; border-radius:6px; text-align:center; color:white; font-size:0.85em;'>"
                f"{genre}</div>",
                unsafe_allow_html=True,
            )


# --- Main Execution ---
if "books_df" not in st.session_state:
    st.session_state.books_df = None

if st.session_state.books_df is None:
    fetch_now = st.button("Fetch Books from Fable")
    if fetch_now:
        with st.spinner("Fetching books..."):
            df = fetch_all_books()
            st.session_state.books_df = df
            st.rerun()
    else:
        st.info("Click the button above to fetch books from Fable.")
else:
    df = st.session_state.books_df

    # --- Sidebar filters ---
    st.sidebar.header("Filters and Sorting")

    search = st.sidebar.text_input("Search title or author")
    selected_genres = st.sidebar.multiselect("Filter by genre", sorted(df["genre"].unique()))
    sort_by = st.sidebar.selectbox(
        "Sort by", ["title", "author", "published_date", "finished_reading"]
    )
    ascending = st.sidebar.checkbox("Ascending", value=True)

    # --- Filtering logic ---
    filtered_df = df.copy()
    if search:
        filtered_df = filtered_df[
            filtered_df["title"].str.contains(search, case=False, na=False) |
            filtered_df["author"].str.contains(search, case=False, na=False)
        ]
    if selected_genres:
        filtered_df = filtered_df[filtered_df["genre"].isin(selected_genres)]

    # --- Sorting logic ---
    if sort_by == "finished_reading":
        filtered_df = filtered_df.sort_values(
            by="finished_datetime", ascending=ascending, na_position="last"
        )
    else:
        if sort_by in filtered_df.columns:
            filtered_df = filtered_df.sort_values(by=sort_by, ascending=ascending)

    # --- Genre colors (consistent) ---
    unique_genres = sorted(filtered_df["genre"].unique())
    genre_colors = {g: color_from_genre(g) for g in unique_genres}

    # --- Genre legend ---
    display_genre_legend(genre_colors)

    # --- View Toggle ---
    view_mode = st.radio("Select View Mode", ["Table View", "Gallery View"], horizontal=True)

    if view_mode == "Table View":
        def highlight_genre(row):
            color = genre_colors.get(row["genre"], "#444444")
            return [f"background-color: {color}20" for _ in row]

        st.subheader("Book Data")
        styled_df = (
            filtered_df[
                ["title", "author", "genre", "pages", "published_date", "started_reading", "finished_reading"]
            ]
            .style.apply(highlight_genre, axis=1)
        )
        st.dataframe(styled_df, use_container_width=True)

        # --- Download Excel ---
        excel_data = df_to_excel_with_colors(filtered_df, genre_colors)
        st.download_button(
            "Download Excel",
            data=excel_data,
            file_name="fable_books.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    elif view_mode == "Gallery View":
        st.subheader("Book Gallery")
        for _, row in filtered_df.iterrows():
            with st.container():
                cols = st.columns([1, 3])
                with cols[0]:
                    if row["cover_image"]:
                        st.image(row["cover_image"], width=110)
                with cols[1]:
                    title = row.get("title", "Untitled")
                    url = row.get("book_url", "")
                    st.markdown(f"**[{title}]({url})**")

                    if row.get("author"):
                        st.write(f"by {row['author']}")

                    genre_color = genre_colors.get(row["genre"], "#888888")
                    st.markdown(
                        f'<span style="background-color:{genre_color}; color:white; '
                        f'padding:3px 8px; border-radius:6px; font-size:0.85em;">'
                        f'{row["genre"]}</span>',
                        unsafe_allow_html=True,
                    )

                    if row.get("finished_datetime"):
                        finished_date = row["finished_datetime"].strftime("%B %d, %Y")
                        st.caption(f"Finished: {finished_date}")
                    elif row.get("published_date"):
                        st.caption(f"Published: {row['published_date']}")
